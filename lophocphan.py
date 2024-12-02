from PyQt6.QtWidgets import QMainWindow, QTableWidgetItem, QMessageBox, QFileDialog
from PyQt6 import uic
from database import Database
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

class LopHocPhanWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        # Load giao diện từ file .ui
        uic.loadUi("lophocphan.ui", self)
        
        # Khởi tạo kết nối database
        self.db = Database()
        
        # Thiết lập các kết nối cho nút
        self.setup_connections()

    def setup_connections(self):
        # Kết nối các nút điều hướng
        self.btnSinhVien.clicked.connect(lambda: self.parent().setCurrentIndex(0))
        self.btnDiemDanh.clicked.connect(lambda: self.parent().setCurrentIndex(1))
        self.btnLichSu.clicked.connect(lambda: self.parent().setCurrentIndex(2))
        self.btnTinhDiem.clicked.connect(lambda: self.parent().setCurrentIndex(4))
        self.btnThoat.clicked.connect(self.close)
        
        # Kết nối nút tìm kiếm và xuất Excel
        self.btnTimKiem.clicked.connect(self.search_students)
        self.btnXuatthongtin.clicked.connect(self.export_to_excel)

    def search_students(self):
        # Lấy tên lớp học phần từ ô tìm kiếm
        class_name = self.txtTimKiem.text().strip()
        if not class_name:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng nhập tên lớp học phần!")
            return
            
        # Tìm kiếm sinh viên theo tên lớp học phần
        query = """
            SELECT sv.ma_sv, sv.ho_ten, sv.gioi_tinh, sv.lop
            FROM sinh_vien sv 
            WHERE sv.lop = %s
            ORDER BY sv.ma_sv
        """
        students = self.db.fetch_data(query, (class_name,))
        
        if not students:
            QMessageBox.information(self, "Thông báo", "Không tìm thấy sinh viên nào trong lớp này!")
            self.tblLopHocPhan.setRowCount(0)
            return
        
        # Hiển thị kết quả lên bảng
        self.tblLopHocPhan.setRowCount(0)
        for row_number, student in enumerate(students):
            self.tblLopHocPhan.insertRow(row_number)
            self.tblLopHocPhan.setItem(row_number, 0, QTableWidgetItem(str(student['ma_sv'])))
            self.tblLopHocPhan.setItem(row_number, 1, QTableWidgetItem(student['ho_ten']))
            self.tblLopHocPhan.setItem(row_number, 2, QTableWidgetItem(student['gioi_tinh']))
            self.tblLopHocPhan.setItem(row_number, 3, QTableWidgetItem(student['lop']))

    def export_to_excel(self):
        """Xuất danh sách sinh viên ra file Excel"""
        # Kiểm tra xem có dữ liệu để xuất không
        if self.tblLopHocPhan.rowCount() == 0:
            QMessageBox.warning(self, "Cảnh báo", "Không có dữ liệu để xuất!")
            return

        # Lấy tên lớp học phần
        class_name = self.txtTimKiem.text().strip()
        if not class_name:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng nhập tên lớp học phần!")
            return

        try:
            # Tạo file Excel mới
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sách sinh viên"

            # Thiết lập tiêu đề
            ws['A1'] = "DANH SÁCH SINH VIÊN"
            ws['A2'] = f"Lớp: {class_name}"
            ws['A3'] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"

            # Định dạng tiêu đề
            for cell in [ws['A1'], ws['A2'], ws['A3']]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left')

            # Thêm header cho bảng
            headers = ["STT", "Mã sinh viên", "Họ và tên", "Giới tính", "Lớp"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Thêm dữ liệu
            for row in range(self.tblLopHocPhan.rowCount()):
                # STT
                ws.cell(row=row+6, column=1, value=row+1).alignment = Alignment(horizontal='center')
                # Các cột dữ liệu
                for col in range(4):
                    cell = ws.cell(row=row+6, column=col+2)
                    cell.value = self.tblLopHocPhan.item(row, col).text()
                    cell.alignment = Alignment(horizontal='left')

            # Điều chỉnh độ rộng cột
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Mở hộp thoại lưu file
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "Lưu file Excel",
                f"DanhSach_{class_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                "Excel Files (*.xlsx)"
            )

            if file_path:
                # Lưu file
                wb.save(file_path)
                QMessageBox.information(
                    self,
                    "Thành công",
                    f"Đã xuất danh sách sinh viên ra file:\n{file_path}"
                )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Lỗi",
                f"Không thể xuất file Excel:\n{str(e)}"
            )
